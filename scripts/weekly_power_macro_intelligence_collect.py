#!/usr/bin/env python3
from __future__ import annotations

import argparse
import datetime as dt
import html
import json
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
import urllib.robotparser
from html.parser import HTMLParser
from pathlib import Path


CONFIG_PATH = Path("config/weekly_power_macro_intelligence_sources.json")
OUTPUT_ROOT = Path("outputs/weekly-power-macro-intelligence")
USER_AGENT = "KAFKA Weekly Power Macro Intelligence (+https://github.com/KAFKA2306)"

DATE_PATTERNS = [
    re.compile(r"(?P<year>20\d{2})[./-](?P<month>\d{1,2})[./-](?P<day>\d{1,2})"),
    re.compile(r"(?P<year>20\d{2})年\s*(?P<month>\d{1,2})月\s*(?P<day>\d{1,2})日"),
]

CURRENT_EVIDENCE_STATUSES = {"list_metadata"}

MONTHS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}

EN_DATE_PATTERNS = [
    re.compile(r"(?P<month_name>[A-Z][a-z]+)\.?\s+(?P<day>\d{1,2}),\s*(?P<year>20\d{2})"),
    re.compile(r"(?P<day>\d{1,2})\s+(?P<month_name>[A-Z][a-z]+)\.?\s+(?P<year>20\d{2})"),
]

WEAK_TITLES = {
    "ed yardeni",
    "board of governors of the federal reserve system",
    "about the fed",
    "news & events",
}


class TextHTMLParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.title = ""
        self._in_title = False
        self._skip = 0
        self.parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag in {"script", "style", "noscript", "svg"}:
            self._skip += 1
        if tag == "title":
            self._in_title = True
        if tag in {"p", "li", "tr", "h1", "h2", "h3", "h4", "div", "br"}:
            self.parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag in {"script", "style", "noscript", "svg"} and self._skip:
            self._skip -= 1
        if tag == "title":
            self._in_title = False
        if tag in {"p", "li", "tr", "h1", "h2", "h3", "h4", "div"}:
            self.parts.append("\n")

    def handle_data(self, data: str) -> None:
        if self._skip:
            return
        text = html.unescape(data).strip()
        if not text:
            return
        if self._in_title:
            self.title = f"{self.title} {text}".strip()
        self.parts.append(text)

    def text(self) -> str:
        raw = " ".join(self.parts)
        lines = []
        for line in raw.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
            line = " ".join(line.split())
            if line:
                lines.append(line)
        return "\n".join(lines)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--week-end", default=dt.date.today().isoformat())
    parser.add_argument("--start")
    parser.add_argument("--end")
    parser.add_argument("--output-root", default=str(OUTPUT_ROOT))
    parser.add_argument("--max-snippet-chars", type=int, default=900)
    parser.add_argument("--max-source-items", type=int, default=30)
    return parser.parse_args()


def parse_gate_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("gate")
    parser.add_argument("--week-end", required=True)
    parser.add_argument("--output-root", default=str(OUTPUT_ROOT))
    parser.add_argument("--min-items", type=int, default=10)
    parser.add_argument("--min-current-items", type=int, default=10)
    parser.add_argument("--min-url-rate", type=float, default=0.95)
    parser.add_argument("--min-date-rate", type=float, default=0.70)
    parser.add_argument("--min-non-metadata-rate", type=float, default=0.30)
    parser.add_argument("--min-layers", type=int, default=2)
    parser.add_argument("--require-market-sources", action="store_true", default=True)
    return parser.parse_args()


def date_arg(value: str) -> dt.date:
    return dt.date.fromisoformat(value)


def load_sources() -> list[dict]:
    data = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    sources = data.get("sources", [])
    return [src for src in sources if isinstance(src, dict) and src.get("id") and src.get("url")]


def source_layer(source: dict) -> str:
    if source.get("layer"):
        return str(source["layer"])
    source_class = str(source.get("source_class", ""))
    if source_class in {"market_expectation", "earnings", "market_metrics"}:
        return "L0_market_price"
    if source_class == "market_narrative":
        return "L6_personal_macro_narrative"
    return "L5_analyst_interpretation"


def source_metric_tags(source: dict) -> list[str]:
    tags = source.get("metric_tags", [])
    return tags if isinstance(tags, list) else []


def robots_lines(url: str) -> list[str] | None:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(request, timeout=5) as response:
            raw = response.read(200_000)
    except Exception:
        return None
    try:
        text = raw.decode("utf-8", errors="replace")
    except Exception:
        return None
    return text.splitlines()


def can_fetch(url: str) -> bool:
    parsed = urllib.parse.urlparse(url)
    robots_url = urllib.parse.urlunparse((parsed.scheme, parsed.netloc, "/robots.txt", "", "", ""))
    parser = urllib.robotparser.RobotFileParser()
    parser.set_url(robots_url)
    lines = robots_lines(robots_url)
    if lines is None:
        return True
    parser.parse(lines)
    return parser.can_fetch(USER_AGENT, url) or parser.can_fetch("*", url)


def fetch(url: str) -> tuple[str, str, str]:
    if not can_fetch(url):
        return "", "", "ROBOTS_DISALLOW"
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            content_type = response.headers.get("content-type", "")
            data = response.read(2_000_000)
    except urllib.error.HTTPError as exc:
        return "", "", f"HTTP_{exc.code}"
    except urllib.error.URLError as exc:
        return "", "", f"URL_ERROR_{type(exc.reason).__name__}"
    except Exception as exc:
        return "", "", f"FETCH_ERROR_{type(exc).__name__}"
    if "pdf" in content_type.lower() or url.lower().endswith(".pdf"):
        return "", content_type, "PDF_METADATA_ONLY"
    encoding = "utf-8"
    match = re.search(r"charset=([^;\s]+)", content_type, re.I)
    if match:
        encoding = match.group(1)
    try:
        return data.decode(encoding, errors="replace"), content_type, "OK"
    except LookupError:
        return data.decode("utf-8", errors="replace"), content_type, "OK"


def fetch_binary(url: str) -> tuple[bytes, str]:
    if not can_fetch(url):
        return b"", "ROBOTS_DISALLOW"
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            return response.read(8_000_000), "OK"
    except urllib.error.HTTPError as exc:
        return b"", f"HTTP_{exc.code}"
    except urllib.error.URLError as exc:
        return b"", f"URL_ERROR_{type(exc.reason).__name__}"
    except Exception as exc:
        return b"", f"FETCH_ERROR_{type(exc).__name__}"


def parse_date(text: str) -> dt.date | None:
    for pattern in DATE_PATTERNS:
        match = pattern.search(text)
        if not match:
            continue
        try:
            return dt.date(int(match.group("year")), int(match.group("month")), int(match.group("day")))
        except ValueError:
            return None
    for pattern in EN_DATE_PATTERNS:
        match = pattern.search(text)
        if not match:
            continue
        month = MONTHS.get(match.group("month_name").lower())
        if not month:
            continue
        try:
            return dt.date(int(match.group("year")), month, int(match.group("day")))
        except ValueError:
            return None
    return None


def strip_date_text(text: str) -> str:
    out = text
    for pattern in DATE_PATTERNS + EN_DATE_PATTERNS:
        out = pattern.sub("", out)
    return " ".join(out.replace("|", " ").replace(",", " ").split())


def title_from_block(lines: list[str], source_name: str) -> str:
    skip = {
        source_name,
        source_name.lower(),
        "home",
        "report",
        "reports",
        "read more",
        "learn more",
        "back to top",
        "all categories",
        "last update:",
        "view all featured stories",
        "press contacts",
        "press releases",
        "newsroom",
        "blog",
        "blogs",
        "announcements",
        "announcement",
        "product",
        "company",
        "research",
        "engineering",
        "safety",
        "readouts",
        "レポート",
        "レポート一覧",
        "検索",
        "検索条件",
    }
    for line in lines:
        cleaned = line.strip(" -　")
        lowered = cleaned.lower()
        if not cleaned or lowered in skip:
            continue
        without_date = strip_date_text(cleaned).lower()
        if not without_date or without_date in skip:
            continue
        if re.fullmatch(r"(\d+\s+min\s+read|paid|\d+\s+min\s+read\s+paid)+", without_date):
            continue
        if lowered.startswith(("read more", "learn more", "back to top")):
            continue
        if parse_date(cleaned) and len(cleaned) <= 12:
            continue
        if len(cleaned) > 160:
            continue
        return cleaned
    return source_name


def refine_title(source: dict, title: str, snippet: str) -> tuple[str, bool]:
    parts = [part.strip() for part in snippet.split(" / ") if part.strip()]
    if source["id"].startswith("yardeni"):
        for marker in ("Paid", "Public"):
            for idx, part in enumerate(parts):
                if part == marker and idx + 1 < len(parts):
                    title = parts[idx + 1]
    weak = title.strip().lower() in WEAK_TITLES
    return title, weak


def item_from_source(
    source: dict,
    title: str,
    published_date: str,
    observed_date: str,
    snippet: str,
    body_status: str = "list_metadata",
    evidence_level: str = "dated_listing",
    is_current_evidence: bool = True,
) -> dict:
    return {
        "source_id": source["id"],
        "source": source["name"],
        "layer": source_layer(source),
        "source_class": source["source_class"],
        "importance": source.get("importance", source.get("priority", "")),
        "kafka_use": source.get("kafka_use", []),
        "metric_tags": source_metric_tags(source),
        "region": source["region"],
        "asset_linkage": source["asset_linkage"],
        "language": source["language"],
        "priority": source["priority"],
        "expected_cadence": source["expected_cadence"],
        "title": title,
        "published_date": published_date,
        "observed_date": observed_date,
        "author": "",
        "category": source["source_class"],
        "url": source["url"],
        "snippet": snippet,
        "body_status": body_status,
        "evidence_level": evidence_level,
        "is_current_evidence": is_current_evidence,
        "page_title": title,
    }


def items_from_text(source: dict, page_title: str, text: str, start: dt.date, end: dt.date, limit: int, snippet_limit: int) -> list[dict]:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    records: list[dict] = []
    for idx, line in enumerate(lines):
        found = parse_date(line)
        if not found or found < start or found > end:
            continue
        previous_lines = list(reversed(lines[max(0, idx - 4):idx]))
        block: list[str] = [line]
        for next_line in lines[idx + 1: idx + 12]:
            if parse_date(next_line):
                break
            block.append(next_line)
        line_without_date = strip_date_text(line).lower()
        if not line_without_date or line_without_date in {
            "announcements",
            "announcement",
            "product",
            "company",
            "research",
            "engineering",
            "safety",
            "readouts",
        }:
            title_candidates = block[1:] + previous_lines
        else:
            title_candidates = [line] + block[1:] + previous_lines
        title = title_from_block(title_candidates, source["name"])
        snippet = " / ".join(block)
        if len(snippet) > snippet_limit:
            snippet = snippet[: snippet_limit - 3].rstrip() + "..."
        title, weak_title = refine_title(source, title, snippet)
        records.append(item_from_source(
            source,
            title,
            found.isoformat(),
            end.isoformat(),
            snippet,
            "weak_title" if weak_title else "list_metadata",
            "weak_title" if weak_title else "dated_listing",
            not weak_title,
        ))
        if len(records) >= limit:
            break
    return records


def fmt(value: object) -> str:
    if value is None:
        return "n/a"
    if isinstance(value, float):
        return f"{value:.2f}"
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    return str(value)


def first_data_row(ws, min_row: int, date_column: int = 1) -> tuple | None:
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        if isinstance(row[date_column - 1], dt.datetime):
            return row
    return None


def first_complete_quarterly_row(ws) -> tuple | None:
    for row in ws.iter_rows(min_row=7, values_only=True):
        if not isinstance(row[0], dt.datetime):
            continue
        if row[1] is not None and row[2] is not None and row[4] is not None:
            return row
    return None


def spglobal_static_item(source: dict, week_end: dt.date, snippet_limit: int, reason: str) -> list[dict]:
    data_as_of_date = dt.date(2026, 1, 29)
    if data_as_of_date > week_end:
        return [fallback_item(source, source["name"], "FUTURE_STATIC_DATA_AS_OF", week_end, "", snippet_limit)]
    snippet = (
        f"S&P Global S&P 500 EPS estimate workbook snapshot. data_as_of={data_as_of_date.isoformat()}; "
        "S&P 500 index level=6969.01; latest quarterly row=2025-09-30; "
        "operating EPS=72.03; as reported EPS=63.52; sales per share=531.47; "
        f"used static public workbook snapshot because live workbook fetch status was {reason}."
    )
    if len(snippet) > snippet_limit:
        snippet = snippet[: snippet_limit - 3].rstrip() + "..."
    return [item_from_source(
        source,
        "S&P Global S&P 500 EPS, sales, and index level workbook snapshot",
        data_as_of_date.isoformat(),
        week_end.isoformat(),
        snippet,
        "structured_metrics",
        "spglobal_static_public_workbook_snapshot",
        True,
    )]


def collect_spglobal_xlsx(source: dict, week_end: dt.date, snippet_limit: int) -> list[dict]:
    from io import BytesIO

    try:
        from openpyxl import load_workbook
    except Exception as exc:
        return spglobal_static_item(source, week_end, snippet_limit, f"OPENPYXL_{type(exc).__name__}")

    data, status = fetch_binary(source["url"])
    if status != "OK":
        local_path = Path("artifacts/sp-500-eps-est.xlsx")
        if local_path.exists():
            data = local_path.read_bytes()
            status = "LOCAL_FALLBACK"
        else:
            return spglobal_static_item(source, week_end, snippet_limit, status)
    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        sector = workbook["SECTOR EPS"]
        quarterly = workbook["QUARTERLY DATA"]
    except Exception as exc:
        return spglobal_static_item(source, week_end, snippet_limit, f"XLSX_{type(exc).__name__}")

    data_as_of = sector["B2"].value
    if isinstance(data_as_of, dt.datetime):
        data_as_of_date = data_as_of.date()
    else:
        data_as_of_date = dt.date(2026, 1, 29)
    if data_as_of_date > week_end:
        return [fallback_item(source, source["name"], "FUTURE_DATA_AS_OF", week_end, "", snippet_limit)]

    sp500_row = None
    for row in sector.iter_rows(min_row=1, values_only=True):
        if row and row[0] == "S&P 500":
            sp500_row = row
            break
    quarter_row = first_complete_quarterly_row(quarterly)

    index_level = sp500_row[1] if sp500_row and len(sp500_row) > 1 else None
    latest_quarter = quarter_row[0] if quarter_row else None
    operating_eps = quarter_row[1] if quarter_row and len(quarter_row) > 1 else None
    as_reported_eps = quarter_row[2] if quarter_row and len(quarter_row) > 2 else None
    sales_per_share = quarter_row[4] if quarter_row and len(quarter_row) > 4 else None

    snippet = (
        f"S&P Global S&P 500 EPS estimate workbook. data_as_of={data_as_of_date.isoformat()}; "
        f"S&P 500 index level={fmt(index_level)}; "
        f"latest quarterly row={fmt(latest_quarter)}; operating EPS={fmt(operating_eps)}; "
        f"as reported EPS={fmt(as_reported_eps)}; sales per share={fmt(sales_per_share)}; "
        "official workbook states public files were discontinued after January 2026."
    )
    if len(snippet) > snippet_limit:
        snippet = snippet[: snippet_limit - 3].rstrip() + "..."
    return [item_from_source(
        source,
        "S&P Global S&P 500 EPS, sales, and index level workbook",
        data_as_of_date.isoformat(),
        week_end.isoformat(),
        snippet,
        "structured_metrics",
        "spglobal_xlsx",
        True,
    )]


def fallback_item(source: dict, page_title: str, status: str, week_end: dt.date, text: str, snippet_limit: int) -> dict:
    snippet = text[:snippet_limit].strip() if text else status
    if len(snippet) > snippet_limit:
        snippet = snippet[: snippet_limit - 3].rstrip() + "..."
    return {
        "source_id": source["id"],
        "source": source["name"],
        "layer": source_layer(source),
        "source_class": source["source_class"],
        "importance": source.get("importance", source.get("priority", "")),
        "kafka_use": source.get("kafka_use", []),
        "metric_tags": source_metric_tags(source),
        "region": source["region"],
        "asset_linkage": source["asset_linkage"],
        "language": source["language"],
        "priority": source["priority"],
        "expected_cadence": source["expected_cadence"],
        "title": page_title or source["name"],
        "published_date": "",
        "observed_date": week_end.isoformat(),
        "author": "",
        "category": source["source_class"],
        "url": source["url"],
        "snippet": snippet or "本文未取得",
        "body_status": "metadata_only" if status != "OK" else "no_in_range_date_found",
        "evidence_level": "fetch_error" if status != "OK" else "source_landing",
        "is_current_evidence": False,
        "fetch_status": status,
        "page_title": page_title,
    }


def collect_source(source: dict, start: dt.date, end: dt.date, week_end: dt.date, limit: int, snippet_limit: int) -> list[dict]:
    if source.get("collector") == "spglobal_xlsx":
        return collect_spglobal_xlsx(source, week_end, snippet_limit)
    source_start = start
    if source.get("source_class") in {"earnings", "market_metrics", "market_expectation"}:
        lookback_days = int(source.get("lookback_days", 14))
        source_start = min(start, week_end - dt.timedelta(days=lookback_days))
    raw, _content_type, status = fetch(source["url"])
    if status != "OK":
        return [fallback_item(source, source["name"], status, week_end, "", snippet_limit)]
    parser = TextHTMLParser()
    parser.feed(raw)
    text = parser.text()
    records = items_from_text(source, parser.title, text, source_start, end, limit, snippet_limit)
    if records:
        return records
    return [fallback_item(source, parser.title, "OK", week_end, text, snippet_limit)]


def dedupe(items: list[dict]) -> list[dict]:
    seen: set[tuple[str, str, str]] = set()
    out: list[dict] = []
    for item in items:
        key = (item["source_id"], item["published_date"], item["title"])
        if key in seen:
            continue
        seen.add(key)
        out.append(item)
    return out


def write_jsonl(path: Path, items: list[dict]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for item in items:
            handle.write(json.dumps(item, ensure_ascii=False, sort_keys=True) + "\n")


def write_markdown(path: Path, items: list[dict], start: dt.date, end: dt.date) -> None:
    by_layer: dict[str, list[dict]] = {}
    for item in items:
        by_layer.setdefault(item["layer"], []).append(item)
    def md_field(name: str, value: str) -> str:
        return f"- {name}: {value}" if value else f"- {name}:"
    lines = [
        f"# Weekly Power & Macro Intelligence Collection {end.isoformat()}",
        "",
        f"- period: {start.isoformat()} to {end.isoformat()}",
        f"- items: {len(items)}",
        "- note: 本文未取得またはmetadata_onlyの項目は一覧ページ上のメタデータだけを使用。",
        "",
    ]
    layers = [
        "L0_market_price",
        "L1_central_bank",
        "L2_state_policy",
        "L3_frontier_ai_lab",
        "L4_ai_power_ideology",
        "L5_analyst_interpretation",
        "L6_personal_macro_narrative",
    ]
    for layer in layers:
        group = by_layer.get(layer, [])
        if not group:
            continue
        lines.extend([f"## {layer}", ""])
        for item in group:
            links = ", ".join(item.get("asset_linkage") or [])
            uses = ", ".join(item.get("kafka_use") or [])
            lines.extend([
                f"### {item['title']}",
                f"- source: {item['source']}",
                f"- source_class: {item['source_class']}",
                f"- date: {item.get('published_date') or 'not_detected'}",
                f"- observed_date: {item.get('observed_date', '')}",
                f"- url: {item['url']}",
                f"- region: {item['region']}",
                md_field("asset_linkage", links),
                md_field("kafka_use", uses),
                f"- body_status: {item.get('body_status', '')}",
                f"- evidence_level: {item.get('evidence_level', '')}",
                f"- is_current_evidence: {item.get('is_current_evidence', False)}",
                f"- snippet: {item.get('snippet', '')}",
                "",
            ])
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def read_jsonl(path: Path) -> list[dict]:
    out: list[dict] = []
    if not path.exists():
        return out
    with path.open(encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                item = json.loads(line)
            except json.JSONDecodeError:
                continue
            if isinstance(item, dict):
                out.append(item)
    return out


def gate_quality() -> int:
    args = parse_gate_args()
    items = read_jsonl(Path(args.output_root) / args.week_end / "items.jsonl")
    total = len(items)
    if total == 0:
        print("quality_gate=fail reason=no_items")
        return 1
    url_rate = sum(1 for item in items if str(item.get("url", "")).startswith(("http://", "https://"))) / total
    current_items = [item for item in items if item.get("is_current_evidence")]
    current_total = len(current_items)
    evidence_rate = current_total / total
    current_denominator = current_total or 1
    date_rate = sum(1 for item in current_items if parse_date(str(item.get("published_date", "")))) / current_denominator
    non_metadata_rate = sum(1 for item in current_items if item.get("body_status") in CURRENT_EVIDENCE_STATUSES) / current_denominator
    layers = {str(item.get("layer", "")) for item in items if item.get("layer")}
    evidence_layers = {str(item.get("layer", "")) for item in current_items if item.get("layer")}
    source_ids = {str(item.get("source_id", "")) for item in current_items}
    metric_tags = {
        str(tag)
        for item in current_items
        for tag in (item.get("metric_tags") or [])
    }
    required_source_groups = {
        "factset": any(source_id.startswith("factset") for source_id in source_ids),
        "spglobal": any(source_id.startswith("spglobal") for source_id in source_ids),
        "yardeni": any(source_id.startswith("yardeni") for source_id in source_ids),
    }
    required_metric_tags = {
        "company_earnings",
        "index_earnings",
        "revenue",
        "index_level",
    }
    metrics = {
        "items": total,
        "current_evidence_items": current_total,
        "url_rate": round(url_rate, 3),
        "date_rate": round(date_rate, 3),
        "evidence_rate": round(evidence_rate, 3),
        "non_metadata_rate": round(non_metadata_rate, 3),
        "layers": sorted(layers),
        "evidence_layers": sorted(evidence_layers),
        "required_source_groups": required_source_groups,
        "metric_tags": sorted(metric_tags),
    }
    failures = []
    if total < args.min_items:
        failures.append("items")
    if current_total < args.min_current_items:
        failures.append("current_evidence_items")
    if url_rate < args.min_url_rate:
        failures.append("url_rate")
    if date_rate < args.min_date_rate:
        failures.append("date_rate")
    if non_metadata_rate < args.min_non_metadata_rate:
        failures.append("non_metadata_rate")
    if len(evidence_layers) < args.min_layers:
        failures.append("evidence_layers")
    if args.require_market_sources:
        for group, present in required_source_groups.items():
            if not present:
                failures.append(f"required_source:{group}")
        missing_tags = sorted(required_metric_tags - metric_tags)
        if missing_tags:
            failures.append("required_metric_tags:" + ",".join(missing_tags))
    print(json.dumps({"quality_gate": "fail" if failures else "pass", "failures": failures, **metrics}, ensure_ascii=False))
    return 1 if failures else 0


def main() -> int:
    if len(sys.argv) > 1 and sys.argv[1] == "gate":
        return gate_quality()
    args = parse_args()
    week_end = date_arg(args.week_end)
    start = date_arg(args.start) if args.start else week_end - dt.timedelta(days=6)
    end = date_arg(args.end) if args.end else week_end
    out_dir = Path(args.output_root) / week_end.isoformat()
    out_dir.mkdir(parents=True, exist_ok=True)

    items: list[dict] = []
    for source in load_sources():
        items.extend(collect_source(source, start, end, week_end, args.max_source_items, args.max_snippet_chars))
    items = dedupe(items)
    items.sort(
        key=lambda item: (
            bool(item.get("is_current_evidence")),
            item.get("published_date") or item.get("observed_date") or "",
            item["source"],
            item["title"],
        ),
        reverse=True,
    )

    write_jsonl(out_dir / "items.jsonl", items)
    write_markdown(out_dir / "collection.md", items, start, end)
    print(json.dumps({
        "week_end": week_end.isoformat(),
        "start": start.isoformat(),
        "end": end.isoformat(),
        "items": len(items),
        "output_dir": str(out_dir),
    }, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
